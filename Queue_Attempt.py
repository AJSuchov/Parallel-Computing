from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool, Queue
import time
from copy import copy, deepcopy
from operator import itemgetter


###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster_Super_Large.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

def print_read(d,k):
    for i in d:
        alist = []
    # iterate over all columns
        blist = []
        for j in range(1,max_column+1):
          # get particular cell value    
            cell_obj=sheet.cell(row=i,column=j)
          # print cell value     
          #list.append(cell_obj.value)
            blist.append(cell_obj.value)
        alist.append(blist)
        k.put_nowait(alist)
     # print new line
    

def find_ms_lvl(use_list, ms1,ms2,ms3,ms4):
    for cadet in use_list:
        if cadet[2] == '1':
            ms1.put(cadet)
        elif cadet[2] == '2':
            ms2.put(cadet)
        elif cadet[2] == '3':
            ms3.put(cadet)
        else:
            ms4.put(cadet)

#Hangs here for large lists
def put_lists(queue, use_list):
    for i in use_list:
        queue.put(i)


def put_in_queue(queue, ms_class):
    for i in ms_class:
        queue.put(i)

    
def addtosheet(full_queue):
    while not full_queue.empty():
        print(full_queue.get())

def full_run():
    full_list = []
    full_list2 = []
    ms1_list = []
    ms2_list = []
    ms3_list = []
    ms4_list = []
    main_list = []
###################################################################################    
    ms1_queue = multiprocessing.Queue()
    ms2_queue = multiprocessing.Queue()
    ms3_queue = multiprocessing.Queue()
    ms4_queue = multiprocessing.Queue()

    full_queue = multiprocessing.Queue()

    quart = int(max_row/4)
    quart1 = multiprocessing.Process(target=print_read, args=(range(2,quart),full_queue))
    quart2 = multiprocessing.Process(target=print_read, args=(range(quart,quart*2),full_queue))
    quart3 = multiprocessing.Process(target=print_read, args=(range(quart*2,quart*3),full_queue))
    quart4 = multiprocessing.Process(target=print_read, args=(range(quart*3,max_row+1),full_queue))
    
    quart1.start()
    quart2.start()
    quart3.start()
    quart4.start()

    quart1.join()
    quart2.join()
    quart3.join()
    quart4.join()
###############################################################################
    while not full_queue.empty():
        full_list.extend(full_queue.get())
    
    ms_lvl = multiprocessing.Process(target=find_ms_lvl, args=(full_list,ms1_queue,ms2_queue,ms3_queue,ms4_queue))

    ms_lvl.start()
    ms_lvl.join()
###############################################################################
    while not ms1_queue.empty():
        ms1_list.append(ms1_queue.get())

    while not ms2_queue.empty():
        ms2_list.append(ms2_queue.get())

    while not ms3_queue.empty():
        ms3_list.append(ms3_queue.get())

    while not ms4_queue.empty():
        ms4_list.append(ms4_queue.get())
    
    ms1_list = sorted(ms1_list)
    ms2_list = sorted(ms2_list)
    ms3_list = sorted(ms3_list)
    ms4_list = sorted(ms4_list)
    
###############################################################################
    full_list2.extend(ms1_list)
    full_list2.extend(ms2_list)
    full_list2.extend(ms3_list)
    full_list2.extend(ms4_list)

    

    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    sheet2.append(header)

    for i in full_list2:
        sheet2.append(i)
    

    
#############################################################################################################

if __name__ == '__main__':
    starttime = time.time()

    full_run()

    wb2.save(filepath2)
    print('That took {} seconds'.format(time.time() - starttime))
