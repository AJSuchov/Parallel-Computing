from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool
import time
from copy import copy, deepcopy
from operator import itemgetter

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocess examples\\Absence_Roster.xlsx"
#filepath="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Absence_Roster.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocess examples\\Sorted_Roster_Blank.xlsx"
#filepath2="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

global sheet2

sheet2=wb2.active

###########################################################################################

list = [] #go back an name this main list
global list_for_1
global list_for_2
global list_for_3
global list_for_4
global mainlist

list_for_1 = []
list_for_2 = []
list_for_3 = []
list_for_4 = []
mainlist = []

def print_read1(k):
    alist = []
    time.sleep(0.5)
    # iterate over all columns
    for j in range(1,max_column+1):
          # get particular cell value    
        cell_obj=sheet.cell(row=k,column=j)
          # print cell value     
          #list.append(cell_obj.value)
        alist.append(cell_obj.value)  
     # print new line
    return alist

def send_ms1_list(k): #k[0] = ['Weisner', 'Jacob', 4, 0]
    time.sleep(0.5)
    list_ms1 = []
    m = []
    if k[2] == 1:
        m = deepcopy(k)
        list_ms1.extend(m)
    return list_ms1

def send_ms2_list(k):
    time.sleep(0.5)
    list_ms2 = []
    m = []
    if k[2] == 2:
        m = deepcopy(k)
        list_ms2.extend(m)
    return list_ms2

def send_ms3_list(k):
    time.sleep(0.5)
    list_ms3 = []
    m = []
    if k[2] == 3:
        m = deepcopy(k)
        list_ms3.extend(m)
    return list_ms3

def send_ms4_list(k):
    time.sleep(0.5)
    list_ms4 = []
    m = []
    if k[2] == 4:
        m = deepcopy(k)
        list_ms4.extend(m)
    return list_ms4

def remove_empty_brackets(testList):
    time.sleep(0.5)
    return testList

def sort_list(testList):
    time.sleep(0.5)
    return testList

def return_append(testList):
    time.sleep(0.5)
    return testList

def append_sheet(list_added):
    time.sleep(0.5)
    return list_added

#############################################################################################################

if __name__ == '__main__':
    starttime = time.time()
    
    p1 = Pool(processes=1)
    p2 = Pool(processes=1)
    p3 = Pool(processes=1)
    p4 = Pool(processes=1)

    data1 = p1.map(print_read1, [i for i in range(2,int(max_row/4))])
    data2 = p2.map(print_read1, [i for i in range(int(max_row/4),int(max_row/4)*2)])
    data3 = p3.map(print_read1, [i for i in range(int(max_row/4)*2,int(max_row/4)*3)])
    data4 = p4.map(print_read1, [i for i in range(int(max_row/4)*3,max_row+1)])

    p1.close()
    p2.close()
    p3.close()
    p4.close()

    
    #See if I can make this happen in parallel as well.
    pext1 = multiprocessing.Process(list.extend(data1))
    pext2 = multiprocessing.Process(list.extend(data2))
    pext3 = multiprocessing.Process(list.extend(data3))
    pext4 = multiprocessing.Process(list.extend(data4))

    pext1.start()
    pext2.start()
    pext3.start()
    pext4.start()

    p5 = Pool(processes=1)
    p6 = Pool(processes=1)
    p7 = Pool(processes=1)
    p8 = Pool(processes=1)

    ms1 = p5.map(send_ms1_list,list) 
    ms2 = p6.map(send_ms2_list,list)
    ms3 = p7.map(send_ms3_list,list)
    ms4 = p8.map(send_ms4_list,list)

    p5.close()
    p6.close()
    p7.close()
    p8.close()

###################################################################################
    p9 = Pool(processes=1)
    p10 = Pool(processes=1)
    p11 = Pool(processes=1)
    p12 = Pool(processes=1)

    #make parallel
    list_for_1 = p9.map(remove_empty_brackets, [x for x in ms1 if x != []])
    list_for_2 = p10.map(remove_empty_brackets, [x for x in ms2 if x != []])
    list_for_3 = p11.map(remove_empty_brackets, [x for x in ms3 if x != []])
    list_for_4 = p12.map(remove_empty_brackets, [x for x in ms4 if x != []])

    p9.close()
    p10.close()
    p11.close()
    p12.close()

#####################################################################################
    p13 = Pool(processes=1)
    p14 = Pool(processes=1)
    p15 = Pool(processes=1)
    p16 = Pool(processes=1)

    list_for_1 = p13.map(sort_list, [sorted(list_for_1, key=itemgetter(0))])
    list_for_2 = p14.map(sort_list, [sorted(list_for_2, key=itemgetter(0))])
    list_for_3 = p15.map(sort_list, [sorted(list_for_3, key=itemgetter(0))])
    list_for_4 = p16.map(sort_list, [sorted(list_for_4, key=itemgetter(0))])
    
    p13.close()
    p14.close()
    p15.close()
    p16.close()

####################################################################################
    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    mainlist.insert(0,header)
    
    pext5 = multiprocessing.Process(mainlist.extend(i for i in list_for_1[0]))
    pext6 = multiprocessing.Process(mainlist.extend(i for i in list_for_2[0]))
    pext7 = multiprocessing.Process(mainlist.extend(i for i in list_for_3[0]))
    pext8 = multiprocessing.Process(mainlist.extend(i for i in list_for_4[0]))

    pext5.start()
    pext6.start()
    pext7.start()
    pext8.start()

##    for i in list_for_1:
##        time.sleep(0.5)
##        mainlist.extend(i)
##
##    for i in list_for_2:
##        time.sleep(0.5)
##        mainlist.extend(i)
##
##    for i in list_for_3:
##        time.sleep(0.5)
##        mainlist.extend(i)
##
##    for i in list_for_4:
##        time.sleep(0.5)
##        mainlist.extend(i)

##################################################################################
    
    with Pool(1) as p:
        for number in range(0,len(mainlist)-1):
            cadet = p.map(append_sheet, [i for i in mainlist])
            sheet2.append(cadet[number])

    wb2.save(filepath2)

    print('That took {} seconds'.format(time.time() - starttime))

    



