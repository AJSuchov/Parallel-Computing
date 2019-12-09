from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool, Queue
import time
from copy import copy, deepcopy
from operator import itemgetter
#import listMod

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster_Super_Large.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

list = [] #go back an name this main list
mainlist = []
list_for_1 = []
list_for_2 = []
list_for_3 = []
list_for_4 = []

def print_read(k,all_list):
    for i in k:
        alist = []
        for j in range(1,max_column+1):
              # get particular cell value    
            cell_obj=sheet.cell(row=i,column=j)
              # print cell value     
              #list.append(cell_obj.value)
            alist.append(cell_obj.value)
        all_list.append(alist)
         # print new line
        #return alist

def append_sheet(cadet):
    return cadet

def return_sort(sorted_list):
    return sorted_list

    
#############################################################################################################

if __name__ == '__main__':
    starttime = time.time()
    manager = multiprocessing.Manager()

    mainlist = []
    all_list = manager.list()
    
    quart = int(max_row/4)
    quart1 = multiprocessing.Process(target=print_read, args=(range(2,quart), all_list))
    quart2 = multiprocessing.Process(target=print_read, args=(range(quart,quart*2), all_list))
    quart3 = multiprocessing.Process(target=print_read, args=(range(quart*2,quart*3), all_list))
    quart4 = multiprocessing.Process(target=print_read, args=(range(quart*3,max_row), all_list))

    quart1.start()
    quart2.start()
    quart3.start()
    quart4.start()

    quart1.join()
    quart2.join()
    quart3.join()
    quart4.join()
#############################################################################################################
    p = Pool(processes=1)
    p1 = Pool(processes=1)
    p2 = Pool(processes=1)
    p3 = Pool(processes=1)
    
    ms1 = p.map(return_sort, [i for i in all_list if i[2] == '1'])
    ms2 = p1.map(return_sort, [i for i in all_list if i[2] == '2'])
    ms3 = p2.map(return_sort, [i for i in all_list if i[2] == '3'])
    ms4 = p3.map(return_sort, [i for i in all_list if i[2] == '4'])

    p.close()
    p1.close()
    p2.close()
    p3.close()
#############################################################################################################
    ms1 = sorted(ms1)
    ms2 = sorted(ms2)
    ms3 = sorted(ms3)
    ms4 = sorted(ms4)

    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    sheet2.append(header)
#############################################################################################################
    p4 = Pool(processes=1)
    p5 = Pool(processes=1)
    p6 = Pool(processes=1)
    p7 = Pool(processes=1)

    
    for number in range(0,len(ms1)):
        cadet = p4.map(append_sheet,[i for i in ms1])
        sheet2.append(cadet[number])

    for number in range(0,len(ms2)):
        cadet = p5.map(append_sheet,[i for i in ms2])
        sheet2.append(cadet[number])

    for number in range(0,len(ms3)):
        cadet = p6.map(append_sheet,[i for i in ms3])
        sheet2.append(cadet[number])

    for number in range(0,len(ms4)):
        cadet = p7.map(append_sheet,[i for i in ms4])
        sheet2.append(cadet[number])

    p4.close()
    p5.close()
    p6.close()
    p7.close()
#############################################################################################################
    wb2.save(filepath2)
    #print('100 Entries for Attempt2 took {} seconds'.format(time.time() - starttime))
    #print('1000 Entries for Attempt2 took {} seconds'.format(time.time() - starttime))
    print('10000 Entries for Attempt2 took {} seconds'.format(time.time() - starttime))
