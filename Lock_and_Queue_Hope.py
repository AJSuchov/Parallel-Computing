from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool, Queue
import time
from copy import copy, deepcopy
from operator import itemgetter


###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"


#filepath="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Absence_Roster.xlsx"
#filepath="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Large_Roster_Test.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

#filepath2="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Sorted_Roster_Blank.xlsx"
filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

def print_read(d,k,lock):
####    lock.acquire()
    i = 0
    lock.acquire()
    for i in d:
        alist = []
    # iterate over all columns
        blist = []
        for j in range(1,max_column+1):
          # get particular cell value    
            cell_obj=sheet.cell(row=i,column=j)
          # print cell value     
          #list.append(cell_obj.value)
            blist.append(cell_obj.value)
        alist.append(blist)
     # print new line
    #print(alist)
        k.put(alist)
        #print(alist, ' ', i, ' ', 'Putting value in to queue')
        i = i + 1
    lock.release()

def find_ms_lvl(do_num, use_queue, ms1,ms2,ms3,ms4,lock):
    lock.acquire()
    i = 0
    for i in do_num:
        cadet = []
        cadet.extend(use_queue.get())
        for c in cadet:
            #print(c, ' ', i, ' ', 'Getting from queue')
            i = i + 1
            if c[2] == '1':
                ms1.put(c)
            elif c[2] == '2':
                ms2.put(c)
            elif c[2] == '3':
                ms3.put(c)
            else:
                ms4.put(c)
    lock.release()




###########################################################################################
if __name__ == '__main__':
    starttime = time.time()

    full_list = []
    full_list2 = []
    ms1_list = []
    ms2_list = []
    ms3_list = []
    ms4_list = []
    main_list = []
    
    ms1_queue = multiprocessing.Queue()
    ms2_queue = multiprocessing.Queue()
    ms3_queue = multiprocessing.Queue()
    ms4_queue = multiprocessing.Queue()

    full_queue = multiprocessing.Queue()

    lock = multiprocessing.Lock()
    lock2 = multiprocessing.Lock()

    quart = int(max_row/4)
    quart1 = multiprocessing.Process(target=print_read, args=(range(2,max_row),full_queue,lock))
    ms_lvl = multiprocessing.Process(target=find_ms_lvl, args=(range(2,max_row),full_queue,ms1_queue,ms2_queue,ms3_queue,ms4_queue,lock2))

    quart1.start()
    
    ms_lvl.start()
    

    
    quart1.join()
    ms_lvl.join()

###########################################################################################    
    while not ms1_queue.empty():
        ms1_list.append(ms1_queue.get())

    while not ms2_queue.empty():
        ms2_list.append(ms2_queue.get())

    while not ms3_queue.empty():
        ms3_list.append(ms3_queue.get())

    while not ms4_queue.empty():
        ms4_list.append(ms4_queue.get())

    ms1_list = sorted(ms1_list)
    ms2_list = sorted(ms2_list)
    ms3_list = sorted(ms3_list)
    ms4_list = sorted(ms4_list)
###########################################################################################
    
    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    sheet2.append(header)

    for i in ms1_list:
        sheet2.append(i)

    for i in ms2_list:
        sheet2.append(i)

    for i in ms3_list:
        sheet2.append(i)

    for i in ms4_list:
        sheet2.append(i)
###########################################################################################
        
    wb2.save(filepath2)
    print('100 Entries for Lock_and_Queue_Hope took {} seconds'.format(time.time() - starttime))
