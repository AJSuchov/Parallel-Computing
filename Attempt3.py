from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool, Queue
import time
from copy import copy, deepcopy
from operator import itemgetter
#import listMod

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocess examples\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster_Super_Large.xlsx"



# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

#filepath2="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Sorted_Roster_Blank.xlsx"
filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

list = [] #go back an name this main list
mainlist = []
list_for_1 = []
list_for_2 = []
list_for_3 = []
list_for_4 = []

def print_read(k):
    alist = []
    # iterate over all columns
    for j in range(1,max_column+1):
          # get particular cell value    
        cell_obj=sheet.cell(row=k,column=j)
          # print cell value     
          #list.append(cell_obj.value)
        alist.append(cell_obj.value)
     # print new line
    return alist


def return_sort(sorted_list):
    return sorted_list

def append_sheet(ms):
    return ms
    
    
#############################################################################################################

if __name__ == '__main__':
    starttime = time.time()

    main_list = []
    p = Pool(processes=4)
    all_list = p.map(print_read, [i for i in range(2,max_row+1)])
    p.close()
#############################################################################################################
    p = Pool(processes=1)
    p1 = Pool(processes=1)
    p2 = Pool(processes=1)
    p3 = Pool(processes=1)
    
    ms1 = p.map(return_sort, [i for i in all_list if i[2] == '1'])
    ms2 = p1.map(return_sort, [i for i in all_list if i[2] == '2'])
    ms3 = p2.map(return_sort, [i for i in all_list if i[2] == '3'])
    ms4 = p3.map(return_sort, [i for i in all_list if i[2] == '4'])

    p.close()
    p1.close()
    p2.close()
    p3.close()
#############################################################################################################
    ms1 = sorted(ms1)
    ms2 = sorted(ms2)
    ms3 = sorted(ms3)
    ms4 = sorted(ms4)

    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    sheet2.append(header)
#############################################################################################################    
    p4 = Pool(processes=1)
    p5 = Pool(processes=1)
    p6 = Pool(processes=1)
    p7 = Pool(processes=1)

    
    for number in range(0,len(ms1)):
        cadet = p4.map(append_sheet,[i for i in ms1])
        sheet2.append(cadet[number])

    for number in range(0,len(ms2)):
        cadet = p5.map(append_sheet,[i for i in ms2])
        sheet2.append(cadet[number])

    for number in range(0,len(ms3)):
        cadet = p6.map(append_sheet,[i for i in ms3])
        sheet2.append(cadet[number])

    for number in range(0,len(ms4)):
        cadet = p7.map(append_sheet,[i for i in ms4])
        sheet2.append(cadet[number])

    p4.close()
    p5.close()
    p6.close()
    p7.close()
#############################################################################################################    
    
    wb2.save(filepath2)
    #print('100 Entries for Attempt3 took {} seconds'.format(time.time() - starttime))
    #print('1000 Entries for Attempt3 took {} seconds'.format(time.time() - starttime))
    print('10000 Entries for Attempt3 took {} seconds'.format(time.time() - starttime))


