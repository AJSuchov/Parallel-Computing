from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import multiprocessing
from multiprocessing import Pool, Queue
import time
from copy import copy, deepcopy
from operator import itemgetter
import listMod

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocess examples\\Absence_Roster.xlsx"
filepath="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Absence_Roster.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

filepath2="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Sorted_Roster_Blank.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

list = [] #go back an name this main list
mainlist = []
list_for_1 = []
list_for_2 = []
list_for_3 = []
list_for_4 = []

def print_read(k):
    alist = []
    #time.sleep(0.5)
    # iterate over all columns
    for j in range(1,max_column+1):
          # get particular cell value    
        cell_obj=sheet.cell(row=k,column=j)
          # print cell value     
          #list.append(cell_obj.value)
        alist.append(cell_obj.value)  
     # print new line
    return alist

def append_sheet(cadet):
    sheet2.append(cadet)

def return_sort(sorted_list):
    return sorted_list

def sort_class(big_list):
    p = Pool(processes=1)
    p1 = Pool(processes=1)
    p2 = Pool(processes=1)
    p3 = Pool(processes=1)
    
    ms1 = p.map(return_sort, [i for i in big_list if i[2] == 1])
    ms2 = p1.map(return_sort, [i for i in big_list if i[2] == 2])
    ms3 = p2.map(return_sort, [i for i in big_list if i[2] == 3])
    ms4 = p3.map(return_sort, [i for i in big_list if i[2] == 4])

    p.close()
    p1.close()
    p2.close()
    p3.close()

    ms1 = sorted(ms1)
    ms2 = sorted(ms2)
    ms3 = sorted(ms3)
    ms4 = sorted(ms4)

    mainlist.extend(ms1)
    mainlist.extend(ms2)
    mainlist.extend(ms3)
    mainlist.extend(ms4)

    p4 = Pool(processes=1)
    p4.map(append_sheet, [i for i in mainlist])
    p4.close()
    
    
def ms_Start():
    p = Pool(processes=4)
    all_list = p.map(print_read, [i for i in range(2,max_row+1)])
    p.close()
    sort_class(all_list)
    
#############################################################################################################

if __name__ == '__main__':
    ms_Start()

    wb2.save(filepath2)




