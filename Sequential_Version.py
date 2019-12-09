from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import time
from copy import copy, deepcopy
from operator import itemgetter

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"
filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster_Super_Large.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column


############################################################################################
########### Appending another excel sheet with sorted values ###############################
######## This is to open that new workbook which will be used latter #######################

filepath2="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Sorted_Roster_Blank_Sequential.xlsx"
#filepath2="C:\\Users\\suchovaj\\Desktop\\Parallel Python\\Sorted_Roster_Blank_Sequential.xlsx"

wb2=load_workbook(filepath2)

sheet2=wb2.active

###########################################################################################

list = [] #go back an name this main list
mainlist = []
list_for_1 = []
list_for_2 = []
list_for_3 = []
list_for_4 = []

def print_read1(start,end):
    alist = []
    
    for i in range(start,end):
        #time.sleep(0.5)
    # iterate over all columns
        blist = []
        for j in range(1,max_column+1):
          # get particular cell value    
            cell_obj=sheet.cell(row=i,column=j)
          # print cell value     
          #list.append(cell_obj.value)
            blist.append(cell_obj.value)
        alist.append(blist)
     # print new line
    return alist

def find_ms_lvl(use_list):
    for cadet in use_list:
        #time.sleep(0.5)
        if cadet[2] == 1:
            list_for_1.append(cadet)
        elif cadet[2] == 2:
            list_for_2.append(cadet)
        elif cadet[2] == 3:
            list_for_3.append(cadet)
        else:
            list_for_4.append(cadet)

def extend_new_main(sorted_list):
    for i in sorted_list:
        #time.sleep(0.5)
        mainlist.extend(i)

if __name__ == '__main__':
    starttime = time.time()
    list = print_read1(2,max_row+1)
    
    find_ms_lvl(list)

    list_for_1 = [sorted(list_for_1, key=itemgetter(0))]
    list_for_2 = [sorted(list_for_2, key=itemgetter(0))]
    list_for_3 = [sorted(list_for_3, key=itemgetter(0))]
    list_for_4 = [sorted(list_for_4, key=itemgetter(0))]

    extend_new_main(list_for_1)
    extend_new_main(list_for_2)
    extend_new_main(list_for_3)
    extend_new_main(list_for_4)

    header = ['Last Name', 'First Name', 'MS Level', 'Absences']
    mainlist.insert(0,header)
    
    for cadet in mainlist:
        #time.sleep(0.5)
        sheet2.append(cadet)

    wb2.save(filepath2)
    #print('100 Entries for Sequential took {} seconds'.format(time.time() - starttime))
    #print('1000 Entries for Sequential took {} seconds'.format(time.time() - starttime))
    print('10000 Entries for Sequential took {} seconds'.format(time.time() - starttime))
    
